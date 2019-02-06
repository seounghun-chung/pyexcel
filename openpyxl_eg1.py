import re
import openpyxl

import time

class _openpyxl(object):
    def __init__(self, filename, sheetname):
        self.wb = openpyxl.load_workbook(filename = filename, read_only=True, data_only=True)
        self.sheet = self.wb[sheetname]

    def search(self, txt, cells, exactly = False):
        ''' 
        search specific text in cell 
        cells var can have 'A1:D20' or Cell dictionary (sheet['A1:D20'])
        ''' 
            
        if type(cells) == str:
            ''' Columnar access is disabled in read-only mode because of the way the data is stored in the XML source. '''
            regix = re.compile("[a-zA-Z]{1,}:[a-zA-Z]{1,}")
            o = regix.search(cells)
            if o is not None:
                cells = cells.replace(':', '1:') + str(self.sheet.max_row)  # temperary patch for selecting all range
            cells = self.sheet[cells]

        # make search function (regix or absolute compare)
        if exactly is True:
            _searchfunc = lambda cell, _find : (str(cell.value) == str(_find))
        else:
            regix = re.compile(txt)        
            _searchfunc = lambda cell, _find : (regix.search(str(cell.value)) != None)
            
        # duplicated loop search
        _search = list()
        for row in cells:
            for v in row:
                if (v.value is None):   # Null cell
                    continue             
                elif (_searchfunc(v,txt) is True):  # found
                    _search.append(v)
                else:   # not found
                    pass
        return _search
        
    def cell(self, row, col):
        ''' dont use to select range. It has long latency '''
        return self.sheet.cell(row,col).value
        
    def autoselect(self, row, col, offsetrow=0, offsetcol=0):
        ''' 
        select range with valid data range with last valid row/col
        offset 1 means itself. e.g.) (10,'D',1,1) select D10:D10 not D10:E11
        '''
        
        ''' firstly, select arbitary sheet cells. After that, we shrink cell range '''
        col = openpyxl.utils.cell.column_index_from_string(col) if type(col) == str else col
        endrow = self.sheet.max_row if (offsetrow == 0) or (row + offsetrow - 1 >= self.sheet.max_row) else row + offsetrow - 1 # offset 1 means itself, so need to minus 1
        endcol = self.sheet.max_column if (offsetcol == 0) or (col + offsetcol - 1 >= self.sheet.max_column)  else col + offsetcol - 1 # offset 1 means itself, so need to minus 1
        startcell = openpyxl.utils.cell.get_column_letter(col) + str(row)        
        endcell = openpyxl.utils.cell.get_column_letter(endcol) + str(endrow)        
        cells = self.sheet[startcell + ':' + endcell]

        ''' find not last NULL cell '''
        lastrow = 0
        for row in cells:
            if (row[0].value != None):  
                lastrow = row
            else:
                for v in lastrow:
                    if (v.value != None):
                        endcell = v
                    else:
                        break
                break
                
        ''' re-range cells '''
        endcell = endcell if type(endcell) != str else self.sheet[endcell]  # if don't find NULL cell, we select ragne with max offset
        cells = self.sheet[startcell + ':' + endcell.coordinate]   
        
        for row in cells:
            for v in row:
                if v.value == None:
                    raise RuntimeError('Range %s has NULL cell' % (startcell + ':' + endcell))
                
        print('autoselect : Select Range %s' % (startcell + ':' + endcell.coordinate))
        return cells

    def pprint(self, cells):
        for row in cells:
            for v in row:
                print(v.value, end='\t')
            print()
            
def mytest1():        
    c = _openpyxl('weather.xlsx', '시트1')
    a = c.autoselect(18,'d',1,10)
    print(a)
    
def mytest2():
    ''' auto select example '''
    c = _openpyxl('example.xlsx', 'Sheet1')
    search = c.search('SQ1', 'G:I',exactly=True)
    cells = c.autoselect(search[0].row, search[0].column, 100, 100)
    print(cells)

def mytest3():
    ''' auto select example '''
    c = _openpyxl('example.xlsx', 'Sheet1')
    search = c.search('PARAM', 'A:P',exactly=False)
    print(search)
    cells = c.autoselect(search[0].row, search[0].column, 100, 100)
    c.pprint(cells)    

mytest3()    
#a = c.search('SQ', a)
#print(a)
#o = c.search('구름많음', 'A2:M10000', exactly = False)
#for cell in o:
#    print(cell.row, cell.column, cell.value)
#    print(c.cell(cell.row, cell.column))
#    print(sheet_ranges.cell(cell.row,cell.column).value)
#print(sheet_ranges.cell(5,2).value)

